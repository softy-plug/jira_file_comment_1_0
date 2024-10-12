import os

os.system("pip install jira")
os.system("pip install openpyxl")

# import libraries
import openpyxl
from openpyxl import load_workbook
from jira import JIRA

input("Для начала программы нажмите Enter")

# Load the Excel workbook
exl_path = r"jira.xlsx"
exl = load_workbook(exl_path)  # Запись пути к таблице

# Работа с листом data
sheet_data = exl["data"]
login_exl = sheet_data.cell(row=2, column=1).value
password_exl = sheet_data.cell(row=2, column=2).value

# Работа с листом jira
sheet_jr = exl["jr"]

# Connect to JIRA server
jira = JIRA(server='https://jr.synergy.ru', basic_auth=(login_exl, password_exl))

# Iterate over all rows in the specified column
for row in range(2, sheet_jr.max_row + 1):  # Start from row 2
    # Read issue key, comment, and file path from the current row
    jr_exl = sheet_jr.cell(row=row, column=1).value
    jr_comment = sheet_jr.cell(row=row, column=2).value
    file_name = sheet_jr.cell(row=row, column=3).value  # Assuming the file path is in the third column

    # JIRA issue key
    issue_key = jr_exl

    # Add comment to the issue
    jira.add_comment(issue_key, jr_comment)

    # Attach the file if it exists
    if os.path.exists(file_name):
        jira.add_attachment(issue=issue_key, attachment=file_name)
        print(f"Файл '{file_name}' прикреплён к задаче {issue_key}.")
    else:
        print(f"Файл '{file_name}' не найден. Задача {issue_key} не была обновлена.")

input("Отправка сообщений в заявки завершена. Нажмите Enter для закрытия окна")

# softy_plug
